# -*- coding: utf-8 -*-
"""
Created on Sat Sep  5 09:18:58 2020

@author: Miki
"""

import os
import SimpleITK as sitk
import numpy as np
from openpyxl import Workbook
import scipy.misc

gt_path = r'C:\Users\Miki\Desktop\SegNetCMR-master\Data\Test\Labels\Sunnybrook_Part3'
prediction_path = r'C:\Users\Miki\Desktop\SegNetCMR-master\Output\infer_test'
gt_list = []
mask_list = []
post_list = []
pre_list = []
pre_average = 0
post_average = 0
row = 1

wb = Workbook()
ws = wb.active
ws['A1'] = "image"
ws['B1'] = "dice score"


#ground turth path
for root, dir, files in os.walk(gt_path):
    for file in files:
       break
   
num = len(files)

for i in range(num):
  
    row = row + 1
    
    ws['A'+str(row)] = files[i]
    
    gt_route = os.path.join(gt_path, files[i])
    pre_route = os.path.join(prediction_path, files[i])
    
    gt = scipy.misc.imread(gt_route)
    pre = scipy.misc.imread(pre_route)
    gt = gt[...,0][...,None]/255
    pre = pre[...,0][...,None]/255
    
    add_label_output = gt + pre
    two_add_gt_pre = np.sum(add_label_output==2)
    one_in_gt= np.sum(gt)
    one_in_pre= np.sum(pre)
    under_val = one_in_gt+one_in_pre #-two_add_gt_pre
    pre_ans = 2*two_add_gt_pre/under_val
    pre_average = pre_average + pre_ans
    #d_pre[pre_list[i]] = pre_ans
    ws['B'+str(row)] = pre_ans
    
    
row = row + 1    
pre_average = pre_average/num
ws['B'+str(row)] = pre_average

print('average dice score: {pre_average}'.format(pre_average=pre_average))
wb.save(r'C:\Users\Miki\Desktop\SegNetCMR-master/test_analysis.xlsx')
print('finish output...')
